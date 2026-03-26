"""
PolypSense3D — Clinical subset for Stage 2b distillation training.

Reference: Zhang et al., "PolypSense3D: A Multi-Source Benchmark Dataset for
           Depth-Aware Polyp Size Measurement in Endoscopy"
           NeurIPS 2025  https://doi.org/10.7910/DVN/K13H89

Layout after extraction (root = <dest>/PolypSense3D):
    root/Clinical-Dataset-For-PolypSense3D/
        calibrations/
            camera.xlsx             per-lens intrinsics + distortion params
            video_info.xlsx         sequence → lens_id mapping
        polyp_frame/
            sequenceXXX/
                polypY/
                    frame_NNNN.jpg  1157×1006 RGB, barrel-distorted (k1≈-0.4)

Content:
    51 calibrated sequences out of 57 (sequences 052–057 lack calibration — skipped)
    ~11,645 frames of real colonoscopy footage
    No GT depth, no poses → single-frame distillation only

Distortion:
    All lenses have k1≈-0.4 (barrel). Barrel correction is safe and improves
    structural clarity. Tangential distortion is negligible (p1=p2=0) for all lenses.
"""

from __future__ import annotations

from pathlib import Path

import cv2
import numpy as np
import torch
from PIL import Image
from torchvision import transforms

from endo_da3.data.base import EndoDepthDataset

_CLINICAL_SUBDIR = "Clinical-Dataset-For-PolypSense3D"


def _load_calibrations(cal_dir: Path) -> dict[str, dict]:
    """
    Returns mapping: sequence_name → dict(fx, fy, cx, cy, dist)
    where dist = np.array([k1, k2, 0, 0, k3]) (OpenCV radial-only format).
    Sequences absent from video_info.xlsx are excluded.
    """
    try:
        import openpyxl
    except ImportError as e:
        raise ImportError("openpyxl is required: pip install openpyxl") from e

    wb_cam = openpyxl.load_workbook(cal_dir / "camera.xlsx")
    cam_params: dict[int, dict] = {}
    for row in wb_cam.active.iter_rows(min_row=2, values_only=True):
        cam_id, fx, fy, cx, cy, _iw, _ih, k1, k2, k3, _p1, _p2 = row
        cam_params[int(cam_id)] = dict(
            fx=float(fx), fy=float(fy), cx=float(cx), cy=float(cy),
            dist=np.array([float(k1), float(k2), 0.0, 0.0, float(k3)],
                          dtype=np.float64),
        )

    wb_info = openpyxl.load_workbook(cal_dir / "video_info.xlsx")
    seq_cal: dict[str, dict] = {}
    for row in wb_info.active.iter_rows(min_row=2, values_only=True):
        _idx, seq_name, cam_id = row
        if int(cam_id) in cam_params:
            seq_cal[str(seq_name)] = cam_params[int(cam_id)]

    return seq_cal


class PolypSense3DClinicalDataset(EndoDepthDataset):
    """
    PolypSense3D clinical subset — real colonoscopy, no GT depth or poses.

    Each sample is a single barrel-undistorted frame returned as a 1-frame
    sequence compatible with EndoDA3's (B, S, 3, H, W) input format.
    Intended for teacher-student distillation (DA3-BASE → Endo-DA3).

    Args:
        root        : PolypSense3D root (contains Clinical-Dataset-For-PolypSense3D/).
        img_size    : Output square resolution (default 336).
        undistort   : Apply barrel-distortion correction (default True).
    """

    name = "polypsense3d_clinical"

    def __init__(
        self,
        root: str | Path,
        img_size: int = 336,
        undistort: bool = True,
    ):
        self.img_size  = img_size
        self.undistort = undistort

        clinical_root = Path(root) / _CLINICAL_SUBDIR
        seq_cal = _load_calibrations(clinical_root / "calibrations")

        self._tf = transforms.Compose([
            transforms.Resize(img_size, interpolation=transforms.InterpolationMode.BICUBIC),
            transforms.CenterCrop(img_size),
            transforms.ToTensor(),
            transforms.Normalize(mean=[0.485, 0.456, 0.406],
                                 std=[0.229, 0.224, 0.225]),
        ])

        self._samples: list[tuple[Path, np.ndarray, np.ndarray]] = []
        frames_root = clinical_root / "polyp_frame"
        for seq_dir in sorted(frames_root.iterdir()):
            if not seq_dir.is_dir():
                continue
            cal = seq_cal.get(seq_dir.name)
            if cal is None:
                continue  # no calibration — skip
            K = np.array([[cal["fx"],       0.0, cal["cx"]],
                          [      0.0, cal["fy"], cal["cy"]],
                          [      0.0,       0.0,      1.0 ]], dtype=np.float64)
            for polyp_dir in sorted(seq_dir.iterdir()):
                if not polyp_dir.is_dir():
                    continue
                for frame_path in sorted(polyp_dir.glob("frame_*.jpg")):
                    self._samples.append((frame_path, K, cal["dist"]))

    def __len__(self) -> int:
        return len(self._samples)

    def __getitem__(self, idx: int) -> dict:
        path, K, dist = self._samples[idx]

        img = np.array(Image.open(path).convert("RGB"))
        if self.undistort:
            img = cv2.undistort(img, K, dist)

        tensor = self._tf(Image.fromarray(img))   # (3, H, W)

        return {
            "images": tensor.unsqueeze(0),         # (1, 3, H, W)
            "depths": torch.zeros(1, self.img_size, self.img_size),
        }
